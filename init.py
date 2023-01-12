import os
from dotenv import load_dotenv, find_dotenv
from playwright.sync_api import sync_playwright
from openpyxl.styles import PatternFill
from openpyxl import Workbook

load_dotenv(find_dotenv())

class Collect:
  def __init__(self, page):
    self.page = page
    self.data = {}
    self.page.locator('xpath=/html/body/div[1]/div[1]/div[1]/div/div[2]/div[4]/div[1]/span').click()
    self.page.locator('xpath=/html/body/div[1]/div[1]/div[1]/div/div[2]/div[4]/div[2]/div/div/div[1]/div[1]/div/div/div/div/div/div/div/div/div[1]/div/div/div[1]/div[1]/div/a').click()
    self.page.locator('xpath=/html/body/div[1]/div[1]/div[1]/div/div[3]/div/div/div/div[1]/div[1]/div/div/div[3]/div/div/div/div[1]/div/div/div[1]/div/div/div/div/div/div/a[2]').click()
  def set_relationship_status(self):
    self.data["A1"] = 'status de relacionamento'
    self.data["A2"] = self.page.locator('xpath=/html/body/div[1]/div[1]/div[1]/div/div[3]/div/div/div/div[1]/div[1]/div/div/div[4]/div/div/div/div[1]/div/div/div/div/div[2]/div/div/div/div[6]/div/div/div[1]/div/div[2]/div/span').text_content()
  def set_sexual_gender(self):
    self.page.locator('xpath=/html/body/div[1]/div[1]/div[1]/div/div[3]/div/div/div/div[1]/div[1]/div/div/div[4]/div/div/div/div[1]/div/div/div/div/div[1]/div[5]/a').click()
    self.data["B1"] = 'gênero sexual'
    self.data["B2"] = self.page.locator('xpath=/html/body/div[1]/div[1]/div[1]/div/div[3]/div/div/div/div[1]/div[1]/div/div/div[4]/div/div/div/div[1]/div/div/div/div/div[2]/div/div/div/div[3]/div[2]/div/div/div[1]/div/div[2]/div/div[1]/div/div/div[1]/span').text_content()
  def set_date_birthday(self):
    self.data["C1"] = 'data de aniversário'
    self.data["C2"] = self.page.locator('xpath=/html/body/div[1]/div[1]/div[1]/div/div[3]/div/div/div/div[1]/div[1]/div/div/div[4]/div/div/div/div[1]/div/div/div/div/div[2]/div/div/div/div[3]/div[3]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div/div[1]/span').text_content()
    

def login(page):
  useremail = os.environ.get('EMA')
  password = os.environ.get('PAS')
  page.locator('xpath=//*[@id="email"]').fill(useremail)
  page.locator('xpath=//*[@id="pass"]').fill(password)
  page.locator('xpath=/html/body/div[1]/div[1]/div[1]/div/div/div/div[2]/div/div[1]/form/div[2]/button').click()

def create_spreadsheet(data):
  wb = Workbook()
  ws = wb.active
  for key, value in data.items():
    ws[key] = value
    if key.find("1")>-1:
      ws[key].fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
  wb.save("data.xlsx")

def init(p):
  url = os.environ.get('URL')
  browser = p.chromium.launch()
  page = browser.new_page()
  page.goto(url)
  login(page)

  collect = Collect(page)
  collect.set_relationship_status()
  collect.set_sexual_gender()
  collect.set_date_birthday()
  create_spreadsheet(collect.data)

  browser.close()

with sync_playwright() as p:
  try:
    init(p)
  except ValueError:
    print(ValueError)



    
