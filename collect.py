import os
from dotenv import load_dotenv, find_dotenv
from playwright.sync_api import sync_playwright
from openpyxl.styles import PatternFill
from openpyxl import Workbook

load_dotenv(find_dotenv())

class Collect:
  def __init__(self, page):
    self.page = page
    self.data = {}
    self.page.locator('xpath=/html/body/div[1]/div[1]/div[1]/div/div[2]/div[4]/div[1]/span').click()
    self.page.locator('xpath=/html/body/div[1]/div[1]/div[1]/div/div[2]/div[4]/div[2]/div/div/div[1]/div[1]/div/div/div/div/div/div/div/div/div[1]/div/div/div[1]/div[1]/div/a').click()
    self.page.locator('xpath=/html/body/div[1]/div[1]/div[1]/div/div[3]/div/div/div/div[1]/div[1]/div/div/div[3]/div/div/div/div[1]/div/div/div[1]/div/div/div/div/div/div/a[2]').click()

  def set_relationship_status(self):
    self.data["A1"] = 'status de relacionamento'
    self.data["A2"] = self.page.locator('xpath=/html/body/div[1]/div[1]/div[1]/div/div[3]/div/div/div/div[1]/div[1]/div/div/div[4]/div/div/div/div[1]/div/div/div/div/div[2]/div/div/div/div[6]/div/div/div[1]/div/div[2]/div/span').text_content()
  def set_sexual_gender(self):
    self.page.locator('xpath=/html/body/div[1]/div[1]/div[1]/div/div[3]/div/div/div/div[1]/div[1]/div/div/div[4]/div/div/div/div[1]/div/div/div/div/div[1]/div[5]/a').click()
    self.data["B1"] = 'gênero sexual'
    self.data["B2"] = self.page.locator('xpath=/html/body/div[1]/div[1]/div[1]/div/div[3]/div/div/div/div[1]/div[1]/div/div/div[4]/div/div/div/div[1]/div/div/div/div/div[2]/div/div/div/div[3]/div[2]/div/div/div[1]/div/div[2]/div/div[1]/div/div/div[1]/span').text_content()
  def set_date_birthday(self):
    self.data["C1"] = 'data de aniversário'
    self.data["C2"] = self.page.locator('xpath=/html/body/div[1]/div[1]/div[1]/div/div[3]/div/div/div/div[1]/div[1]/div/div/div[4]/div/div/div/div[1]/div/div/div/div/div[2]/div/div/div/div[3]/div[3]/div/div/div[1]/div/div[2]/div[1]/div[1]/div/div/div[1]/span').text_content()
  def friends_list(self):
    friends_list = self.page.evaluate("""() => {
      var names = [];
      var docs = document.querySelectorAll("div > div.x9f619.x1n2onr6.x1ja2u2z > div > div > div > div.x78zum5.xdt5ytf.x10cihs4.x1t2pt76.x1n2onr6.x1ja2u2z > div.x78zum5.xdt5ytf.x1t2pt76 > div > div > div.x6s0dn4.x78zum5.xdt5ytf.x193iq5w > div > div > div > div:nth-child(2) > div > div > div > div > div:nth-child(3) > div > div > div.x1iyjqo2.x1pi30zi > div:nth-child(1) > a > span");
      docs.forEach((doc) => {names.push(doc.innerHTML)})
      return names;
    }""")    
    self.data["D1"] = 'amigos'
    for index in range(len(friends_list)):
      self.data["D"+str(index+2)] = friends_list[index]

def login(page, data_user):
  useremail = data_user.get('email')
  password = data_user.get('password')
  page.locator('xpath=//*[@id="email"]').fill(useremail)
  page.locator('xpath=//*[@id="pass"]').fill(password)
  page.locator('xpath=/html/body/div[1]/div[1]/div[1]/div/div/div/div[2]/div/div[1]/form/div[2]/button').click()

def create_spreadsheet(data):
  wb = Workbook()
  ws = wb.active
  for key, value in data.items():
    ws[key] = value
    if key.find("1")>-1:
      ws[key].fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
  wb.save("data.xlsx")

def init(p, data_user):
  url = "https://pt-br.facebook.com/"
  browser = p.chromium.launch(headless=False)
  page = browser.new_page()
  page.goto(url)
  login(page, data_user)
  collect = Collect(page)
  collect.set_relationship_status()
  collect.set_sexual_gender()
  collect.set_date_birthday()
  collect.friends_list()
  create_spreadsheet(collect.data)
  browser.close()



    
